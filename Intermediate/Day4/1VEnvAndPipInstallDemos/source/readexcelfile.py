import openpyxl as excel

filepath = '../Data/students.xlsx'

workbook = excel.load_workbook(filepath)
currentsheet = workbook['Mar2019']



for row in range( 1, currentsheet.max_row+1):
    for col in range(1, currentsheet.max_column+1):
        print(currentsheet.cell(row, col).value, end=' | ')
    
    print()
    
    