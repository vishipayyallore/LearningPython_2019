import openpyxl as excel

filepath = '../Data/students.xlsx'

workbook = excel.load_workbook(filepath)
currentsheet = workbook['Mar2019']



for row in range( 2, currentsheet.max_row+1):
    sum = 0

    for col in range(3, 6):
        sum += int(currentsheet.cell(row, col).value)
    
    total = currentsheet.cell(row, 8)
    total.value = sum

workbook.save(filepath)

    